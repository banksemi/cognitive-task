from psychopy import visual, core, event, gui
from datetime import datetime, timedelta
from psychopy import iohub
import mouse as ms
# Import analyze_log
import random
import openpyxl
import winsound
import threading
import os  
import json
import sys
from PIL import Image



import pyaudio
import numpy as np
import time
####################################################################################
# 액셀 입출력을 위한 클래스
class pyresult:
    value_table = {
        'Digit Span': {
            'Participant ID': ['A',2],
            'Family name': ['B',2],
            'First name': ['C',2],

            'block_span': ['F', 2],
            
            'trial_stimulus': ['B', 5],
            'trial_response': ['C', 5],
            'trial_correct': ['D', 5],
            'trial_score': ['E', 5],
        },
        'Corsi': {
            'Participant ID': ['A',2],
            'Family name': ['B',2],
            'First name': ['C',2],

            'block_span': ['F', 2],
            
            'trial_stimulus': ['B', 5],
            'trial_response': ['C', 5],
            'trial_reaction_time1': ['D', 5],
            'trial_reaction_time2': ['E', 5],
            'trial_correct': ['F', 5],
            'trial_score': ['G', 5],
        },
        'Go, No-Go': {
            'Participant ID': ['A',2],
            'Family name': ['B',2],
            'First name': ['C',2],

            'inhibition': ['D', 2],
            'score': ['E', 2],
            'go_reaction_time': ['H', 2],
            'nogo_reaction_time': ['I', 2],
            
            'trial_stimulus': ['B', 5],
            'trial_response': ['C', 5],
            'trial_correct': ['D', 5],
            'trial_reaction_time': ['E', 5],
        },
        'DCCS': {
            'Participant ID': ['B',2],
            'Family name': ['C',2],
            'First name': ['D',2],

            'trial1_score': ['E', 2],
            'trial1_reaction_time': ['F', 2],
            'trial1_reaction_time_correct': ['G', 2],
            'trial1_each_stimulus': ['C', 5],
            'trial1_each_response': ['D', 5],
            'trial1_each_correct': ['E', 5],
            'trial1_each_time': ['F', 5],
            
            
            'trial2_score': ['H', 2],
            'trial2_reaction_time': ['I', 2],
            'trial2_reaction_time_correct': ['J', 2],
            'trial2_each_stimulus': ['C', 17],
            'trial2_each_response': ['D', 17],
            'trial2_each_correct': ['E', 17],
            'trial2_each_time': ['F', 17],
            
            
            'trial3_score': ['K', 2],
            'trial3_reaction_time': ['L', 2],
            'trial3_reaction_time_correct': ['M', 2],
            'trial3_each_stimulus': ['C', 29],
            'trial3_each_response': ['D', 29],
            'trial3_each_correct': ['E', 29],
            'trial3_each_time': ['F', 29],
        }
    }
    
    def __init__(self, participant_info, test_name, task_type, reset = True):
        self.test_name = test_name
        self.participant_info = participant_info
        self.task_type = task_type
        self.output_path = '../output/%s_%s_%s_CBT.xlsx' % (participant_info['Participant ID'], participant_info['Family name'], participant_info['First name'])
        sheet_name = self.test_name + ' (' + task_type + ')'
        if os.path.isfile(self.output_path):
            self.workbook = openpyxl.load_workbook(self.output_path)

            if reset:
                original = openpyxl.load_workbook('../template/' + '메인AB' + '.xlsx')
                for i in 'ABCDEFGHIJKLMN':
                    for j in range(1,100):
                        self.workbook[sheet_name][i + str(j)] = original[sheet_name][i + str(j)].value
                
                original.close()
        else:
            self.workbook = openpyxl.load_workbook('../template/' + '메인AB' + '.xlsx')
        self.worksheet = self.workbook[sheet_name]
        for i in participant_info:
            self.write(i, participant_info[i])
        
    def write(self, name, value, index=0):
        if name not in self.value_table[self.test_name]:
            return
        position = self.value_table[self.test_name][name]
        if isinstance(value, list):
            value = str(value)
        self.worksheet[position[0] + str(position[1] + index)] = value
    
    def read(self, name, index=0):
        position = self.value_table[self.test_name][name]
        return self.worksheet[position[0] + str(position[1] + index)].value
        
    def save(self, reload=True):
        try:
            self.workbook.save(self.output_path)
        except PermissionError as e:
            myDlg = gui.Dlg(title="오류")
            myDlg.addText('엑셀 파일을 쓸 수 없습니다. (이미 실행중 혹은 권한 오류)')
            ok_data = myDlg.show()  # show dialog and wait for OK or Cancel
            raise e
        self.__init__(self.participant_info, self.test_name, self.task_type, reset= False)
        
    def close(self):
        self.workbook.close()
        
    def __enter__(self):
        return self
        
    def __exit__(self, exc_type, exc_value, traceback):
        self.save(reload=False)
        self.close()

# 객체 및 입력 이벤트 관리를 위한 Common Module 정의
class drawling_object:
    z = 0
    visible = True
    def __init__(self, dobject):
        self.dobject = dobject
        
    def setVisible(self, value):
        self.visible = value
        
    def draw(self):
        self.dobject.draw()
        
    def contains(self, pos):
        return self.dobject.contains(pos)
        
    def update(self):
        return

class PassException(Exception):    # Exception을 상속받아서 새로운 예외를 만듦
    def __init__(self):
        super().__init__('연습 시행 생략')

class window_manager:
    threads = []
    dobjects = []
    input_keys = []
    event_listener_exit = []
    exited = False
    def __init__(self, win):
        self.win = win
        self.reset_state()
        self.mouse=psycopy_mouse(visible=True, win=win)
        t = threading.Thread(target=self.background_sound)
        t.start()

    def background_sound(self):
        p = pyaudio.PyAudio()
        fs = 44100       # sampling rate, Hz, must be integer
        duration = 0.5   # in seconds, may be float
        f = 24000.0        # sine frequency, Hz, may be float

        # generate samples, note conversion to float32 array
        samples = (np.sin(2*np.pi*np.arange(fs*duration)*f/fs)).astype(np.float32)

        # for paFloat32 sample values must be in range [-1.0, 1.0]
        stream = p.open(format=pyaudio.paFloat32,
                        channels=1,
                        rate=fs,
                        output=True)
        while self.exited == False:
            stream.write(0.001*samples)
            time.sleep(duration)
    
    def append(self, dobject):
        self.dobjects.append(dobject)
        
    def remove(self, dobject):
        self.dobjects.remove(dobject)
    
    def getPressKey(self, key):
        for i in self.input_keys:
            if i == key:
                return True
        return False
        

    def exit(self):
        self.exited = True
    def update(self):
        self.mouse.update()
        self.win.update()
        self.input_keys = event.getKeys()
        if self.getPressKey('p'):
            if len(self.state) != 0:
                raise PassException()
        if self.getPressKey('escape'):
            for event_function in self.event_listener_exit:
                event_function()
            self.exited = True
            self.win.close()
            core.quit()
        self.dobjects.sort(key= lambda x: x.z)
        for i in self.dobjects:
            i.update()
            if i.visible:
                i.draw()
            
    def update_wait_key(self, key=None):
        while True:
            self.update()
            if key is None:
                if len(self.input_keys) > 0:
                    return
            else:
                if self.getPressKey(key):
                    return
                
    def update_wait_time(self, seconds=1):
        start_time = datetime.now()
        while (datetime.now() - start_time).total_seconds() < seconds:
            self.update()
            
    def isClickedObject(self, dobject):
        return self.mouse.getClicked() and dobject.contains(self.mouse.getPos())

    def save_state(self):
        self.reset_state()
        for i in self.dobjects:
            self.state.append((i,i.visible, i.z))

    def load_state(self):
        self.dobjects = []
        for i in self.state:
            i[0].visible = i[1]
            i[0].z = i[2]
            self.dobjects.append(i[0])
    def reset_state(self):
        self.state = []

            
# 터치스크린 문제를 해결하기 위해 mouse 객체 재정의
# 기존 window를 이용하는 mouse 객체와, ioHub로 얻은 mouse 객체를 함께 사용
class psycopy_mouse:
    def __init__(self, win, io=None, visible=None, touch_mode = False):
        self.win_mouse=event.Mouse(visible=True,win=win)
        if io is None:
            io = iohub.launchHubServer()
        self.io_mouse = io.devices.mouse
        self.force_position = None
        self.clear()
        
        self.touch_is_working = not touch_mode
        self.visible = True

        if visible is not None:
            self.setVisible(visible)
    def clear(self):
        self.io_mouse.clearEvents() # 이미 쌓여있는 이벤트는 초기화
        self.pressed = False
        self.clicked = False
        self.last_pos = self.getPos()
        
    def setVisible(self, value):
        self.win_mouse.setVisible(value)
        if self.touch_is_working == False and self.visible == False and value == True:
            ms.move(0, 0)
        self.visible = value

    def getClicked(self):
        return self.clicked
        

    def getPos(self):
        if self.force_position is not None:
            return self.force_position
        return self.win_mouse.getPos()
        
    def update(self):
        self.force_position = None # 강제 지정 초기화
        self.clicked = False # 클릭 초기화
        events = self.io_mouse.getEvents(event_type=(iohub.constants.EventConstants.MOUSE_BUTTON_PRESS))
        for e in events:
            print(e)
            self.touch_is_working = True
            if self.pressed == False:
                self.clicked = True
            self.pressed = True
        
        # release 신호 처리
        events = self.io_mouse.getEvents(event_type=(iohub.constants.EventConstants.MOUSE_BUTTON_RELEASE))
        for e in events:
            self.pressed = False
        
        # 터치 스크린 작동이 보장되지 않을 때 마우스 이동을 이용해서 터치 감지
        diff = (self.getPos() - self.last_pos)  * [1920, 1080]
        if self.touch_is_working == False and self.clicked == False:
            if (abs(diff[0]) > 2 or abs(diff[1]) > 2):
                self.clicked = True
                # 이번 프레임 클릭 위치 기억
                self.force_position = self.win_mouse.getPos() 
                # 위치 초기화
                ms.move(0, 0)
                self.last_pos = (0, 0)
        self.last_pos = self.getPos()

            
image_size_cache = {}
def get_image_size(filename):
    global image_size_cache
    if filename in image_size_cache:
        return image_size_cache[filename]
    else:
        image1 = Image.open(filename)
        image_size_cache[filename] = image1.size
        return image1.size

class drawling_image(drawling_object):
    def __init__(self, x, y, image, height = 1):
        global win
        mag1_size = get_image_size(image)
        self.x = x
        self.y = y
        self.width = (mag1_size[0] / mag1_size[1]) * height
        self.height = height
        self.image = visual.ImageStim(win, pos=[x, y], image=image, size=[self.width, height])
        super().__init__(self.image)
        
class drawling_text(drawling_object):
    def __init__(self, x, y, text, color, height = 1):
        global win
        
        self.text = visual.TextStim(win, text=text, color=color, pos=[x, y], height=height)
        super().__init__(self.text)
    def setText(self, text):
        self.text.setText(text)
        
def inputParticipant(title='Participant'):
    with open('../id.txt', 'r') as f:
        json_data = json.load(f)
    title += ' (' + getTaskType() + ')'
    dlg = gui.DlgFromDict(json_data, title=title, order = ['Participant ID', 'Family name', 'First name'])
    if not dlg.OK:
        print ('User Cancelled')
        core.quit()
    id = json_data['Participant ID']
    with open('../id.txt', 'w', encoding='utf-8') as make_file:
        json.dump(json_data, make_file, indent="\t")

    return json_data


def showExplanation(images):
    if not isinstance(images, list):
        images = [images]
    for file_name in images:
        image = drawling_image(0, 0, file_name)
        image.z = 999999
        window.append(image)
        window.update_wait_key('return')
        window.remove(image)
        
win = None
window = None
def initWindow():
    #윈도우
    global win, window
    win = visual.Window([1920, 1080], allowGUI=True, fullscr=True, units='height', color=[255,255,255])
    window = window_manager(win)
    return win, window

def getTaskType():
    if len(sys.argv) == 1:
        return 'A'
    else:
        return sys.argv[1]

####################################################################################